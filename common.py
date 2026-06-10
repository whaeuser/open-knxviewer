"""Helpers shared between server.py (private) and server_public.py (public).

Everything here is stateless: pure functions over a parsed project dict
or a raw .knxproj file on disk.
"""
import io
import logging
import re
import xml.etree.ElementTree as ET
import zipfile

from xknxproject.zip.extractor import extract as knxproj_extract

logger = logging.getLogger(__name__)


# ── ETS security data ─────────────────────────────────────────────────────────


def parse_ets_certificate(raw: str) -> dict:
    """Parse ETS Cloud License certificate format into a dict."""
    fields = {}
    for m in re.finditer(r'(\w+)=(?:"([^"]*)"|([\w+/=]+))', raw):
        fields[m.group(1)] = m.group(2) if m.group(2) is not None else m.group(3)
    return fields


def extract_security_data(tmp_path: str, password: str, project: dict) -> dict:
    """Parse KNX Security data (device keys/passwords, GA keys, ETS cert) from raw project XML."""
    result: dict = {"devices": [], "ga_keys": {}, "ets_certificates": []}
    try:
        with knxproj_extract(tmp_path, password or None) as content:
            f = content.open_project_0()
            xml_str = f.read().decode("utf-8")

        ns_match = re.search(r'xmlns="([^"]+)"', xml_str)
        ns = ns_match.group(1) if ns_match else "http://knx.org/xml/project/21"
        root = ET.fromstring(xml_str)

        # Build raw_address → formatted address map from parsed project
        raw_to_addr: dict[int, str] = {
            ga["raw_address"]: ga["address"]
            for ga in project.get("group_addresses", {}).values()
        }

        # ── Device security — walk topology to reconstruct individual addresses ──
        for area in root.iter(f"{{{ns}}}Area"):
            area_addr = area.get("Address", "0")
            for line in area.iter(f"{{{ns}}}Line"):
                line_addr = line.get("Address", "0")
                for dev in line.iter(f"{{{ns}}}DeviceInstance"):
                    sec = dev.find(f"{{{ns}}}Security")
                    if sec is None:
                        continue
                    dev_addr = dev.get("Address", "0")
                    ia = f"{area_addr}.{line_addr}.{dev_addr}"
                    dev_info = project.get("devices", {}).get(ia, {})
                    ip_cfg = dev.find(f"{{{ns}}}IPConfig")
                    bus_ifaces = []
                    for bi in dev.iter(f"{{{ns}}}BusInterface"):
                        pwd = bi.get("Password")
                        if pwd:
                            bus_ifaces.append(
                                {"ref_id": bi.get("RefId", ""), "password": pwd}
                            )
                    tool_key = sec.get("ToolKey")
                    device_auth_code = sec.get("DeviceAuthenticationCode")
                    device_mgmt_password = sec.get("DeviceManagementPassword")
                    sequence_number = sec.get("SequenceNumber")
                    # Skip devices with only a default SequenceNumber="0" and no actual keys/passwords
                    # (ETS writes <Security SequenceNumber="0"/> to all devices even in non-secure projects)
                    has_keys = (
                        tool_key
                        or device_auth_code
                        or device_mgmt_password
                        or bus_ifaces
                    )
                    has_nonzero_seq = sequence_number not in (None, "0")
                    if not has_keys and not has_nonzero_seq:
                        continue
                    result["devices"].append(
                        {
                            "address": ia,
                            "name": dev_info.get("name") or dev.get("Name") or "",
                            "ip_address": ip_cfg.get("IPAddress")
                            if ip_cfg is not None
                            else None,
                            "mac_address": ip_cfg.get("MACAddress")
                            if ip_cfg is not None
                            else None,
                            "tool_key": tool_key,
                            "device_auth_code": device_auth_code,
                            "device_mgmt_password": device_mgmt_password,
                            "sequence_number": sequence_number,
                            "bus_interfaces": bus_ifaces or None,
                        }
                    )

        # ── GA keys ──────────────────────────────────────────────────────────
        for ga_el in root.iter(f"{{{ns}}}GroupAddress"):
            key = ga_el.get("Key")
            if not key:
                continue
            raw = ga_el.get("Address")
            try:
                raw_int = int(raw) if raw is not None else None
            except ValueError:
                raw_int = None
            formatted = raw_to_addr.get(raw_int, raw or "")
            result["ga_keys"][formatted] = key

        # ── ETS certificates ─────────────────────────────────────────────────
        with zipfile.ZipFile(tmp_path) as zf:
            for name in zf.namelist():
                if name.endswith(".certificate"):
                    raw = zf.read(name).decode("utf-8", errors="replace")
                    cert = parse_ets_certificate(raw)
                    if cert:
                        result["ets_certificates"].append(cert)

    except Exception as exc:
        logger.warning("Security data extraction failed: %s", exc)
    return result


# ── XLSX export ───────────────────────────────────────────────────────────────


def dpt_str(dpt: dict | None) -> str:
    if not dpt or dpt.get("main") is None:
        return ""
    main = dpt["main"]
    sub = dpt.get("sub")
    return f"{main}.{str(sub).zfill(3)}" if sub is not None else str(main)


def flag_str(co: dict) -> str:
    f = co.get("flags") or {}
    out = []
    for key, letter in (("read", "R"), ("write", "W"), ("transmit", "T"),
                        ("update", "U"), ("communication", "C")):
        if f.get(key):
            out.append(letter)
    return "".join(out)


def build_project_xlsx(project: dict) -> bytes:
    """Render the whole project as a multi-sheet XLSX workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    header_align = Alignment(vertical="center")

    def add_sheet(title: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        for col_idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for r in rows:
                v = r[col_idx - 1] if col_idx - 1 < len(r) else ""
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    devs = project.get("devices") or {}

    # 1) Geräte
    dev_rows = [[addr, d.get("name", ""), d.get("manufacturer_name", ""),
                 d.get("order_number") or "", d.get("application") or "",
                 len(d.get("communication_object_ids") or []), d.get("description") or ""]
                for addr, d in devs.items()]
    add_sheet("Geräte",
              ["Adresse", "Name", "Hersteller", "Bestellnr.", "Applikation", "KO-Anzahl", "Beschreibung"],
              dev_rows)

    # 2) Gruppenadressen
    ga_rows = [[ga.get("address", ""), ga.get("name", ""), dpt_str(ga.get("dpt")),
                ga.get("description") or "",
                "; ".join(ga.get("communication_object_ids") or [])]
               for ga in (project.get("group_addresses") or {}).values()]
    add_sheet("Gruppenadressen",
              ["Adresse", "Name", "DPT", "Beschreibung", "Verknüpfte KOs"], ga_rows)

    # 3) Kommunikationsobjekte
    co_rows = []
    for co in (project.get("communication_objects") or {}).values():
        dev_addr = co.get("device_address", "")
        dev = devs.get(dev_addr, {})
        co_rows.append([dev_addr, dev.get("name", ""), co.get("number", ""),
                        co.get("name", ""),
                        dpt_str((co.get("dpts") or [None])[0]),
                        flag_str(co),
                        "; ".join(co.get("group_address_links") or [])])
    add_sheet("Kommunikationsobjekte",
              ["Gerät PA", "Gerät", "KO-Nr.", "Name", "DPT", "Flags", "Gruppenadressen"], co_rows)

    # 4) Funktionen
    fn_rows = []
    for fn in (project.get("functions") or {}).values():
        gas = [v.get("address", "") for v in (fn.get("group_addresses") or {}).values()]
        fn_rows.append([fn.get("identifier", ""), fn.get("name", ""),
                        fn.get("type", ""), "; ".join(gas)])
    add_sheet("Funktionen", ["ID", "Name", "Typ", "Gruppenadressen"], fn_rows)

    # 5) Standorte (flach mit Pfad)
    loc_rows = []

    def walk(node: dict, path: list[str]) -> None:
        for sp in (node.get("spaces") or {}).values():
            here = path + [sp.get("name", "")]
            loc_rows.append([" / ".join(here), sp.get("type", ""),
                             sp.get("usage_text") or "",
                             "; ".join(sp.get("devices") or []),
                             "; ".join(sp.get("functions") or [])])
            walk(sp, here)

    for top in (project.get("locations") or {}).values():
        loc_rows.append([top.get("name", ""), top.get("type", ""),
                         top.get("usage_text") or "",
                         "; ".join(top.get("devices") or []),
                         "; ".join(top.get("functions") or [])])
        walk(top, [top.get("name", "")])
    add_sheet("Standorte", ["Pfad", "Typ", "Nutzung", "Geräte", "Funktionen"], loc_rows)

    # 6) Topologie
    topo_rows = []
    for area_id, area in (project.get("topology") or {}).items():
        for line_id, line in (area.get("lines") or {}).items():
            for dev_addr in (line.get("devices") or []):
                d = devs.get(dev_addr, {})
                topo_rows.append([area_id, area.get("name", ""),
                                  line_id, line.get("name", ""),
                                  dev_addr, d.get("name", "")])
    add_sheet("Topologie",
              ["Bereich", "Bereichsname", "Linie", "Linienname", "Gerät PA", "Gerätename"], topo_rows)

    # Remove the auto-created empty default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_filename(project: dict) -> str:
    project_name = (project.get("info", {}) or {}).get("name") or "knx-projekt"
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", project_name)
    return f"{safe}.xlsx"
