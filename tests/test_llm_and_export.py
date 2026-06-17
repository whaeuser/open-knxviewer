"""Tests for /api/llm/config, /api/export/xlsx and GA-/Bus-Scan guards."""
import io
import json

import server
from routers.llm import LLM_DEFAULT_MODEL


# ---------------------------------------------------------------------------
# LLM config
# ---------------------------------------------------------------------------

async def test_llm_config_defaults(server_client, patched_paths):
    r = await server_client.get("/api/llm/config")
    assert r.status_code == 200
    data = r.json()
    assert data["configured"] is False
    assert data["model"] == LLM_DEFAULT_MODEL


async def test_llm_config_roundtrip(server_client, patched_paths):
    r = await server_client.post(
        "/api/llm/config", json={"api_key": "sk-test", "model": "foo/bar"}
    )
    assert r.status_code == 200
    data = (await server_client.get("/api/llm/config")).json()
    assert data["configured"] is True
    assert data["model"] == "foo/bar"
    # Key must be persisted but never returned via the API
    assert "sk-test" not in json.dumps(data)
    cfg = json.loads((patched_paths / "config.json").read_text())
    assert cfg["openrouter_api_key"] == "sk-test"


async def test_llm_config_empty_model_resets_to_default(server_client, patched_paths):
    await server_client.post("/api/llm/config", json={"model": ""})
    data = (await server_client.get("/api/llm/config")).json()
    assert data["model"] == LLM_DEFAULT_MODEL


async def test_llm_local_model_counts_as_configured(server_client, patched_paths):
    await server_client.post("/api/llm/config", json={"model": "local-model"})
    data = (await server_client.get("/api/llm/config")).json()
    assert data["configured"] is True


async def test_llm_local_url_and_token_roundtrip(server_client, patched_paths):
    r = await server_client.post(
        "/api/llm/config",
        json={"local_url": "http://localhost:10240/v1", "local_token": "sk-local"},
    )
    assert r.status_code == 200
    data = (await server_client.get("/api/llm/config")).json()
    assert data["local_url"] == "http://localhost:10240/v1"
    assert data["local_token_set"] is True
    # Token persisted but never returned via the API
    assert "sk-local" not in json.dumps(data)
    cfg = json.loads((patched_paths / "config.json").read_text())
    assert cfg["local_llm_url"] == "http://localhost:10240/v1"
    assert cfg["local_llm_token"] == "sk-local"


def test_resolve_local_target_uses_config(patched_paths):
    import core
    from routers.llm import _resolve_llm_target

    core.update_config(
        {"local_llm_url": "http://host:9999/v1/", "local_llm_token": "tok"}
    )
    url, headers, model = _resolve_llm_target("local-model", "")
    assert url == "http://host:9999/v1/chat/completions"
    assert headers["Authorization"] == "Bearer tok"
    assert model == "local-model"


async def test_llm_analyze_requires_config(server_client, patched_paths):
    r = await server_client.post("/api/llm/analyze", json={"question": "Was?"})
    assert r.status_code == 400


async def test_llm_analyze_requires_project(server_client, patched_paths):
    await server_client.post("/api/llm/config", json={"api_key": "sk-test"})
    r = await server_client.post("/api/llm/analyze", json={"question": "Was?"})
    assert r.status_code == 400
    assert "Projekt" in r.json()["detail"]


async def test_llm_compare_requires_diff_text(server_client, patched_paths):
    await server_client.post("/api/llm/config", json={"api_key": "sk-test"})
    r = await server_client.post("/api/llm/compare", json={"diff_text": ""})
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

_PROJECT = {
    "info": {"name": "Export Test"},
    "devices": {
        "1.1.1": {
            "name": "Taster",
            "manufacturer_name": "ACME",
            "communication_object_ids": ["co1"],
        }
    },
    "group_addresses": {
        "ga1": {"address": "1/2/3", "name": "Licht", "dpt": {"main": 1, "sub": 1}},
    },
    "communication_objects": {
        "co1": {
            "name": "Schalten",
            "number": 1,
            "device_address": "1.1.1",
            "dpts": [{"main": 1, "sub": 1}],
            "flags": {"read": True, "write": True},
            "group_address_links": ["1/2/3"],
        }
    },
    "functions": {},
    "locations": {},
    "topology": {
        "1": {"name": "Bereich", "lines": {"1": {"name": "Linie", "devices": ["1.1.1"]}}}
    },
}

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _assert_valid_xlsx(content: bytes):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content))
    assert "Geräte" in wb.sheetnames
    assert "Gruppenadressen" in wb.sheetnames
    ws = wb["Gruppenadressen"]
    assert ws.cell(row=2, column=1).value == "1/2/3"


async def test_export_xlsx_requires_project(server_client):
    r = await server_client.get("/api/export/xlsx")
    assert r.status_code == 400


async def test_export_xlsx_private(server_client):
    server.state["project_data"] = _PROJECT
    r = await server_client.get("/api/export/xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == _XLSX_MEDIA_TYPE
    assert "Export_Test.xlsx" in r.headers["content-disposition"]
    _assert_valid_xlsx(r.content)


async def test_export_xlsx_public(public_client):
    r = await public_client.post("/api/export/xlsx", json=_PROJECT)
    assert r.status_code == 200
    assert r.headers["content-type"] == _XLSX_MEDIA_TYPE
    _assert_valid_xlsx(r.content)


# ---------------------------------------------------------------------------
# GA ops / scans — guards without a connection
# ---------------------------------------------------------------------------

async def test_ga_write_requires_connection(server_client):
    r = await server_client.post("/api/ga/write", json={"ga": "1/2/3", "value": "1"})
    assert r.status_code == 503


async def test_ga_write_unknown_dpt_422(server_client):
    server.state["connected"] = True
    r = await server_client.post("/api/ga/write", json={"ga": "1/2/3", "value": "1"})
    assert r.status_code == 422


async def test_ga_read_requires_connection(server_client):
    r = await server_client.post("/api/ga/read", json={"ga": "1/2/3"})
    assert r.status_code == 503


async def test_ga_scan_requires_connection(server_client):
    r = await server_client.post("/api/ga/scan", json={})
    assert r.status_code == 503


async def test_ga_scan_invalid_range_422(server_client):
    server.state["connected"] = True
    r = await server_client.post("/api/ga/scan", json={"start": "kaputt", "end": "1/2/3"})
    assert r.status_code == 422


async def test_bus_scan_requires_connection(server_client):
    r = await server_client.post("/api/bus/scan", json={})
    assert r.status_code == 503


async def test_bus_scan_rejected_via_remote_gateway(server_client):
    server.state["connected"] = True
    server.state["connection_type"] = "remote_gateway"
    r = await server_client.post("/api/bus/scan", json={})
    assert r.status_code == 503
