"""
Public read-only Open-KNXViewer — no bus monitor, no gateway connection.
Safe to expose to the internet.

Run with:
  .venv/bin/uvicorn server_public:app --host 0.0.0.0 --port 8004
"""
import io
import logging
import os
import re
import tempfile
import time
from datetime import datetime
from logging.handlers import TimedRotatingFileHandler
from pathlib import Path

from fastapi import Body, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address
from starlette.middleware.base import BaseHTTPMiddleware

from xknxproject import XKNXProj
from xknxproject.exceptions import InvalidPasswordException, XknxProjectException
from xknxproject.zip.extractor import extract as knxproj_extract

INDEX_HTML = Path(__file__).parent / "index.html"
STATIC_DIR = Path(__file__).parent / "static"
DEMO_PATH = Path(__file__).parent / "demo.knxproj"
ACCESS_LOG = Path(__file__).parent / "logs" / "access_public.log"

MAX_UPLOAD_BYTES = 200 * 1024 * 1024  # 200 MB

# IPs, die nicht ins Access-Log geschrieben werden (z.B. Monitoring)
ACCESS_LOG_SKIP_IPS = {"172.18.0.1"}

# ── Access-Logger ─────────────────────────────────────────────────────────────
ACCESS_LOG.parent.mkdir(exist_ok=True)
_access_handler = TimedRotatingFileHandler(
    ACCESS_LOG, when="midnight", backupCount=30, encoding="utf-8"
)
_access_handler.setFormatter(logging.Formatter("%(message)s"))
access_log = logging.getLogger("access_public")
access_log.setLevel(logging.INFO)
access_log.addHandler(_access_handler)
access_log.propagate = False

limiter = Limiter(key_func=get_remote_address, default_limits=[])
app = FastAPI(title="Open-KNXViewer (Public)", docs_url=None, redoc_url=None)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Content-Security-Policy"] = (
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://unpkg.com; "
            "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://unpkg.com https://fonts.googleapis.com; "
            "img-src 'self' data:; "
            "connect-src 'self'; "
            "font-src 'self' data: https://fonts.gstatic.com; "
            "frame-ancestors https://volt-logik.io https://*.volt-logik.io https://portal.nurdaheim.net https://*.nurdaheim.net;"
        )
        return response


class AccessLogMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        start = time.monotonic()
        response = await call_next(request)
        duration = time.monotonic() - start
        ip = request.headers.get("x-forwarded-for", request.client.host if request.client else "-").split(",")[0].strip()
        if ip in ACCESS_LOG_SKIP_IPS:
            return response
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        access_log.info("%s | %s | %s | %s | %s | %.3fs",
                        ts, ip, request.method, request.url.path,
                        response.status_code, duration)
        return response


app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(AccessLogMiddleware)

if STATIC_DIR.is_dir():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

_demo_cache = None


def _parse_ets_certificate(raw: str) -> dict:
    import re
    fields = {}
    for m in re.finditer(r'(\w+)=(?:"([^"]*)"|([\w+/=]+))', raw):
        fields[m.group(1)] = m.group(2) if m.group(2) is not None else m.group(3)
    return fields


def _extract_security_data(tmp_path: str, password: str, project: dict) -> dict:
    import re
    import xml.etree.ElementTree as ET
    import zipfile
    result: dict = {"devices": [], "ga_keys": {}, "ets_certificates": []}
    try:
        with knxproj_extract(tmp_path, password or None) as content:
            f = content.open_project_0()
            xml_str = f.read().decode("utf-8")

        ns_match = re.search(r'xmlns="([^"]+)"', xml_str)
        ns = ns_match.group(1) if ns_match else "http://knx.org/xml/project/21"
        root = ET.fromstring(xml_str)

        raw_to_addr: dict[int, str] = {
            ga["raw_address"]: ga["address"]
            for ga in project.get("group_addresses", {}).values()
        }

        for area in root.iter(f"{{{ns}}}Area"):
            area_addr = area.get("Address", "0")
            for line in area.iter(f"{{{ns}}}Line"):
                line_addr = line.get("Address", "0")
                for dev in line.iter(f"{{{ns}}}DeviceInstance"):
                    sec = dev.find(f"{{{ns}}}Security")
                    if sec is None:
                        continue
                    dev_addr = dev.get("Address", "0")
                    ia = f"{area_addr}.{line_addr}.{dev_addr}"
                    dev_info = project.get("devices", {}).get(ia, {})
                    ip_cfg = dev.find(f"{{{ns}}}IPConfig")
                    bus_ifaces = []
                    for bi in dev.iter(f"{{{ns}}}BusInterface"):
                        pwd = bi.get("Password")
                        if pwd:
                            bus_ifaces.append({"ref_id": bi.get("RefId", ""), "password": pwd})
                    tool_key = sec.get("ToolKey")
                    device_auth_code = sec.get("DeviceAuthenticationCode")
                    device_mgmt_password = sec.get("DeviceManagementPassword")
                    sequence_number = sec.get("SequenceNumber")
                    # Skip devices with only a default SequenceNumber="0" and no actual keys/passwords
                    # (ETS writes <Security SequenceNumber="0"/> to all devices even in non-secure projects)
                    has_keys = tool_key or device_auth_code or device_mgmt_password or bus_ifaces
                    has_nonzero_seq = sequence_number not in (None, "0")
                    if not has_keys and not has_nonzero_seq:
                        continue
                    result["devices"].append({
                        "address": ia,
                        "name": dev_info.get("name") or dev.get("Name") or "",
                        "ip_address": ip_cfg.get("IPAddress") if ip_cfg is not None else None,
                        "mac_address": ip_cfg.get("MACAddress") if ip_cfg is not None else None,
                        "tool_key": tool_key,
                        "device_auth_code": device_auth_code,
                        "device_mgmt_password": device_mgmt_password,
                        "sequence_number": sequence_number,
                        "bus_interfaces": bus_ifaces or None,
                    })

        for ga_el in root.iter(f"{{{ns}}}GroupAddress"):
            key = ga_el.get("Key")
            if not key:
                continue
            raw = ga_el.get("Address")
            try:
                raw_int = int(raw) if raw is not None else None
            except ValueError:
                raw_int = None
            formatted = raw_to_addr.get(raw_int, raw or "")
            result["ga_keys"][formatted] = key

        with zipfile.ZipFile(tmp_path) as zf:
            for name in zf.namelist():
                if name.endswith(".certificate"):
                    raw = zf.read(name).decode("utf-8", errors="replace")
                    cert = _parse_ets_certificate(raw)
                    if cert:
                        result["ets_certificates"].append(cert)

    except Exception:
        pass
    return result


@app.get("/.well-known/appspecific/com.chrome.devtools.json", include_in_schema=False)
async def chrome_devtools():
    return {}


@app.get("/api/mode")
def get_mode(request: Request):
    host = request.headers.get("x-forwarded-host") or request.headers.get("host", "")
    if "volt-logik" in host:
        theme = "voltlogik"
        branding = "voltlogik"
    else:
        theme = os.environ.get("OPENKNXVIEWER_THEME", "default").strip().lower()
        if theme not in ("default", "voltlogik"):
            theme = "default"
        branding = None
    return {"public": True, "default_theme": theme, "branding": branding}


@app.get("/")
async def root():
    return FileResponse(INDEX_HTML)


@app.get("/api/demo/available")
def demo_available():
    return {"available": DEMO_PATH.exists()}


@app.get("/api/demo")
async def get_demo():
    global _demo_cache
    if not DEMO_PATH.exists():
        raise HTTPException(status_code=404, detail="Demo nicht verfügbar")
    if _demo_cache is None:
        try:
            _demo_cache = XKNXProj(path=str(DEMO_PATH)).parse()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Demo konnte nicht geladen werden: {exc}") from exc
    return JSONResponse(content=_demo_cache)


# ── XLSX export (client supplies project) ────────────────────────────────────

def _dpt_str(dpt: dict | None) -> str:
    if not dpt or dpt.get("main") is None:
        return ""
    main = dpt["main"]
    sub = dpt.get("sub")
    return f"{main}.{str(sub).zfill(3)}" if sub is not None else str(main)


def _flag_str(co: dict) -> str:
    f = co.get("flags") or {}
    out = []
    for key, letter in (("read", "R"), ("write", "W"), ("transmit", "T"),
                        ("update", "U"), ("communication", "C")):
        if f.get(key):
            out.append(letter)
    return "".join(out)


def _build_project_xlsx(project: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    header_align = Alignment(vertical="center")

    def add_sheet(title: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for r in rows:
                v = r[col_idx - 1] if col_idx - 1 < len(r) else ""
                if v is None:
                    continue
                ln = len(str(v))
                if ln > max_len:
                    max_len = ln
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    dev_rows = [[addr, d.get("name", ""), d.get("manufacturer_name", ""),
                 d.get("order_number") or "", d.get("application") or "",
                 len(d.get("communication_object_ids") or []), d.get("description") or ""]
                for addr, d in (project.get("devices") or {}).items()]
    add_sheet("Geräte",
              ["Adresse", "Name", "Hersteller", "Bestellnr.", "Applikation", "KO-Anzahl", "Beschreibung"],
              dev_rows)

    ga_rows = [[ga.get("address", ""), ga.get("name", ""), _dpt_str(ga.get("dpt")),
                ga.get("description") or "",
                "; ".join(ga.get("communication_object_ids") or [])]
               for ga in (project.get("group_addresses") or {}).values()]
    add_sheet("Gruppenadressen",
              ["Adresse", "Name", "DPT", "Beschreibung", "Verknüpfte KOs"], ga_rows)

    co_rows = []
    devs = project.get("devices") or {}
    for co in (project.get("communication_objects") or {}).values():
        dev_addr = co.get("device_address", "")
        dev = devs.get(dev_addr, {})
        co_rows.append([dev_addr, dev.get("name", ""), co.get("number", ""),
                        co.get("name", ""),
                        _dpt_str((co.get("dpts") or [None])[0]),
                        _flag_str(co),
                        "; ".join(co.get("group_address_links") or [])])
    add_sheet("Kommunikationsobjekte",
              ["Gerät PA", "Gerät", "KO-Nr.", "Name", "DPT", "Flags", "Gruppenadressen"], co_rows)

    fn_rows = []
    for fn in (project.get("functions") or {}).values():
        gas = [v.get("address", "") for v in (fn.get("group_addresses") or {}).values()]
        fn_rows.append([fn.get("identifier", ""), fn.get("name", ""),
                        fn.get("type", ""), "; ".join(gas)])
    add_sheet("Funktionen", ["ID", "Name", "Typ", "Gruppenadressen"], fn_rows)

    loc_rows = []
    def walk(node: dict, path: list[str]) -> None:
        for sp in (node.get("spaces") or {}).values():
            here = path + [sp.get("name", "")]
            loc_rows.append([" / ".join(here), sp.get("type", ""),
                             sp.get("usage_text") or "",
                             "; ".join(sp.get("devices") or []),
                             "; ".join(sp.get("functions") or [])])
            walk(sp, here)
    for top in (project.get("locations") or {}).values():
        loc_rows.append([top.get("name", ""), top.get("type", ""),
                         top.get("usage_text") or "",
                         "; ".join(top.get("devices") or []),
                         "; ".join(top.get("functions") or [])])
        walk(top, [top.get("name", "")])
    add_sheet("Standorte", ["Pfad", "Typ", "Nutzung", "Geräte", "Funktionen"], loc_rows)

    topo_rows = []
    for area_id, area in (project.get("topology") or {}).items():
        for line_id, line in (area.get("lines") or {}).items():
            for dev_addr in (line.get("devices") or []):
                d = devs.get(dev_addr, {})
                topo_rows.append([area_id, area.get("name", ""),
                                  line_id, line.get("name", ""),
                                  dev_addr, d.get("name", "")])
    add_sheet("Topologie",
              ["Bereich", "Bereichsname", "Linie", "Linienname", "Gerät PA", "Gerätename"], topo_rows)

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/api/export/xlsx")
@limiter.limit("10/minute")
async def export_xlsx_public(request: Request, project: dict = Body(...)):
    if not project or not isinstance(project, dict):
        raise HTTPException(status_code=400, detail="Kein Projekt übergeben")
    xlsx_bytes = _build_project_xlsx(project)
    project_name = (project.get("info", {}) or {}).get("name") or "knx-projekt"
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", project_name)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}.xlsx"'},
    )


@app.post("/api/parse")
@limiter.limit("5/minute")
async def parse_project(
    request: Request,
    file: UploadFile = File(...),
    password: str = Form(default=""),
    language: str = Form(default="de-DE"),
):
    # 4. Dateiendung validieren
    filename = (file.filename or "").strip()
    if not filename.lower().endswith(".knxproj"):
        raise HTTPException(status_code=400, detail="Nur .knxproj-Dateien sind erlaubt.")

    # 1. Dateigröße begrenzen
    data = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Datei zu groß (max. 50 MB).")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".knxproj", mode="wb") as tmp:
        tmp.write(data)
        tmp_path = tmp.name

    try:
        kwargs: dict = {"path": tmp_path}
        if password:
            kwargs["password"] = password
        if language:
            kwargs["language"] = language

        project = XKNXProj(**kwargs).parse()
        project["_security"] = _extract_security_data(tmp_path, password, project)
        return JSONResponse(content=project)
    except InvalidPasswordException as exc:
        raise HTTPException(status_code=422, detail=f"Invalid password: {exc}") from exc
    except XknxProjectException as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Parsing failed: {exc}") from exc
    finally:
        os.unlink(tmp_path)
